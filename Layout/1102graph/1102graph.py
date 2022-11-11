import matplotlib.pyplot as plt
import math
import json
import openpyxl
import networkx as nx
import community.community_louvain as community_louvain


def read_node(data):
    NODE_ID = []
    NODE_NAME = []
    NODE_X = []
    NODE_Y = []
    REC = {}

    ANGLE = []
    DEPTH = []

    keys = list(data.keys())
    for i in range(len(keys)):
        end = keys[i].find("-")

        node_id = keys[i][:end]
        NODE_ID.append(node_id)

        node_name = keys[i][end + 1:]
        NODE_NAME.append(node_name)
    NODE_ID = list(map(int, NODE_ID))

    for i in range(len(keys)):
        angle = float(data[keys[i]].get("angel")[0])
        depth = float(data[keys[i]].get("depth"))
        rec = float(data[keys[i]].get("rec"))
        ANGLE.append(angle)
        DEPTH.append(depth)
        REC[NODE_ID[i]] = rec

    for i in range(len(ANGLE)):

        x = DEPTH[i] * math.sin(math.radians(ANGLE[i]))
        if ANGLE[i] >= 0:
            y = DEPTH[i] * math.cos(math.radians(ANGLE[i]))
        else:
            y = -DEPTH[i] * math.cos(math.radians(ANGLE[i]))
        NODE_X.append(x)
        NODE_Y.append(y)

    NODE_X = list(map(float, NODE_X))
    NODE_Y = list(map(float, NODE_Y))

    # plot test
    # plt.scatter(NODE_X, NODE_Y)
    # plt.show()
    return NODE_ID, NODE_NAME, NODE_X, NODE_Y, REC


def init_graph(NODE_ID):
    # normal graph: 指定node和edge，随机布局
    G = nx.Graph()
    G.add_nodes_from(range(len(NODE_ID)))
    # print(nx.nodes(G))

    # random powerlaw，指定node与每个node上连接的edge的数量，其他全部随机
    # G = nx.powerlaw_cluster_graph(len(NODE_ID),1,0,1)
    # print(nx.nodes(G))

    mapping = {}
    for i in range(len(NODE_ID)):
        mapping[i] = NODE_ID[i]
    # print(mapping)
    G = nx.relabel_nodes(G, mapping)
    # print(nx.nodes(G))
    return G


def add_edge(G):
    # No relation
    edge = []
    # # Tree
    # edge = [(25, 20), (25, 21), (25, 65), (20, 21), (20, 65)]
    # # Sky
    # edge = [(16, 0), (16, 1), (16, 49), (16, 51), (16, 77)]
    # # Tree+sky
    # edge = [(25, 20), (25, 21), (25, 65), (20, 21), (20, 65), (16, 0), (16, 1), (16, 49), (16, 51), (16, 77)]

    # read edge from excel
    wb = openpyxl.load_workbook("1102relation.xlsx")
    sheet = wb.active
    for i in range(2, sheet.max_row+1):
        source = sheet.cell(row=i, column=1)
        source = int(source.value)
        target = sheet.cell(row=i, column=4)
        target = int(target.value)
        weight = sheet.cell(row=i, column=6)
        weight = float(weight.value)
        relation = (source, target, {"weight": weight})
        edge.append(relation)
    # print(edge)

    G.add_edges_from(edge)
    return G


def get_position(G, NODE_ID, NODE_X, NODE_Y):
    pos = {}
    for i in range(len(nx.nodes(G))):
        # position = dict(node_id[i]=(node_x[i], node_y[i]))
        pos[NODE_ID[i]] = [NODE_X[i], NODE_Y[i]]
    # print(pos)
    return pos


def scale(NODE_ID, pos):
    x_max = 0
    y_max = 0
    for i in NODE_ID:
        x = pos[i][0]
        y = pos[i][1]
        if abs(x) >= x_max:
            x_max = abs(x)
        if abs(y) >= y_max:
            y_max = abs(y)
    # print("max", x_max, y_max)
    if x_max/24 >= y_max/19:
        scale = x_max/24
    else:
        scale = y_max/19
    # print(scale)

    for i in NODE_ID:
        [pos[i][0], pos[i][1]] = [pos[i][0]/scale, pos[i][1]/scale]
    # print(pos)
    return pos


def layout(pos, G, n):
    m = 0
    while m < n:
        pos = scale(NODE_ID, pos)

        # 如果node多，用spring；如果edge多，用planar
        # spring_layout: 无重叠调整node, 可以指定初始位置与不移动的node, 调整iteration可以一定程度上限制移动位置
        # node密集可以增加iteration的次数
        pos = nx.spring_layout(G, pos=pos, k=12, fixed=[NODE_ID[-1]], iterations=2)


        # planar_layout: 无edge交叉的布局方式
        # pos = nx.planar_layout(G)

        # print(pos)
        pos = scale(NODE_ID, pos)
        m = m+1
    return pos


def draw(pos, G, NODE_ID, r):
    nx.draw(G, pos=pos, node_color=range(len(NODE_ID)), node_size=r, with_labels=True, cmap=plt.cm.Blues)
    plt.xlim(-26, 26)
    plt.ylim(-21, 21)
    plt.show()


def draw_o(G, pos):
    O = G.copy()
    nx.draw(O, pos=pos, with_labels=True)
    plt.show()


def grouping(G):
    cluster = community_louvain.best_partition(G, weight="weight")
    groups = {}
    for i, v in cluster.items():
        groups[v] = [i] if v not in groups.keys() else groups[v] + [i]
    groups = dict(sorted(groups.items()))
    # print(groups)
    # print(cluster)
    return cluster, groups


def new_data(data, NODE_ID, NODE_NAME, cluster):
    keys = list(data.keys())
    for i in range(len(keys)):
        data[NODE_ID[i]] = data.pop(keys[i])
        data[NODE_ID[i]]["name"] = NODE_NAME[i]
        group_in_order = list(cluster.values())
        data[NODE_ID[i]]["group"] = group_in_order[i]


def layer(data, cluster, NODE_ID):
    # layer 1
    LAYER1_NODE = []
    group_number = max(cluster.values())
    for i in range(0, group_number+1):
        rec_max = -1
        node = -1
        for j in range(0, len(data)):
            if int(data[NODE_ID[j]]["group"]) == i:
                # group_now = i
                rec = data[NODE_ID[j]]["rec"]
                if rec > rec_max:
                    rec_max = rec
                    node = NODE_ID[j]
        LAYER1_NODE.append(node)
    # print(LAYER1_NODE)

    keys = list(data.keys())
    data_layer1 = data
    for i in range(0, len(keys)):
        if keys[i] not in LAYER1_NODE:
            del data_layer1[keys[i]]
    return data_layer1



if __name__ == "__main__":
    f = open('ang_dep_all.json')
    data = json.load(f)
    # print(data)
    NODE_ID, NODE_NAME, NODE_X, NODE_Y, REC = read_node(data)
    G = init_graph(NODE_ID)
    G = add_edge(G)
    pos_origin = get_position(G, NODE_ID, NODE_X, NODE_Y)
    # pos = layout(pos_origin, G, 5)
    cluster, groups = grouping(G)
    # print(groups)
    new_data(data, NODE_ID, NODE_NAME, cluster)
    # print(data)


    data_layer1 = layer(data, cluster, NODE_ID)
    print(data_layer1)
    # NODE_ID, NODE_NAME, NODE_X, NODE_Y, REC = read_node(data_layer1)
    # G = init_graph(NODE_ID)
    # pos_origin = get_position(G, NODE_ID, NODE_X, NODE_Y)
    # pos = layout(pos_origin, G, 5)



    # draw(pos, G, NODE_ID, 50)
    #
    # draw_o(G, pos_origin)






