import openpyxl
data0=[]
data1=[]
data2=[]
with open("rel.txt", "r") as f:
    for lines in f:
        data = lines.split(",")
        # print(data)
        data0.append(data[0])
        data1.append(data[1])
        data2.append(data[2])

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1, value="Source")
sheet.cell(row=1, column=2, value="relation")
sheet.cell(row=1, column=3, value="target")
sheet.cell(row=1, column=4, value="tensor")
for i in range(len(data0)):
    sheet.cell(row=i+2, column=1, value=data0[i])
    sheet.cell(row=i+2, column=2, value=data1[i])
    sheet.cell(row=i+2, column=3, value=data2[i])

wb.save("1031relation.xlsx")
